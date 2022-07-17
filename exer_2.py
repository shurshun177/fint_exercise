import openpyxl


def append_excel(file_name, results, results_1):
    excel_file = openpyxl.load_workbook(fr'{file_name}')
    excel_file.create_sheet(title='Results', index=3)

    new_sheet = excel_file['Results']
    new_sheet.cell(row=1, column=1).value = 'Table 1'
    new_sheet.cell(row=2, column=1).value = 'Name'
    new_sheet.cell(row=2, column=2).value = 'Result'
    new_sheet.cell(row=2, column=3).value = 'Year'
    keys = ['int_1', 'int_2', 'final']
    i = 3
    for key in keys:
        new_sheet.cell(row=i, column=1).value = key
        new_sheet.cell(row=i, column=2).value = results[key]
        new_sheet.cell(row=i, column=3).value = results['year']
        i += 1
    for key in keys:
        new_sheet.cell(row=i, column=1).value = key
        new_sheet.cell(row=i, column=2).value = results_1[key]
        new_sheet.cell(row=i, column=3).value = results_1['year']
        i += 1


    excel_file.save(filename=file_name)


def proc_excel():
    excel_file = openpyxl.load_workbook(r'fintastic/exercise2.xlsx')
    values_sheet = excel_file['Values']
    formula_sheet = excel_file['Formula']
    val_dict = {}
    val_dict_1 = {}
    for i in range(3, values_sheet.max_row):
        if values_sheet.cell(row=i, column=1).value:
            val_dict.setdefault(
                values_sheet.cell(row=i, column=1).value,
                values_sheet.cell(row=i, column=2).value
            )
            val_dict.setdefault('year', values_sheet.cell(row=i, column=3).value)
            if val_dict.get(values_sheet.cell(row=i, column=1).value):
                val_dict_1[values_sheet.cell(row=i, column=1).value] = \
                    values_sheet.cell(row=i, column=2).value
            if val_dict.get('year'):
                if values_sheet.cell(row=i, column=3).value:
                    val_dict_1['year'] = values_sheet.cell(row=i, column=3).value

    print(val_dict)
    print(val_dict_1)
    formula_dict = {
        formula_sheet.cell(row=i, column=1).value:
        formula_sheet.cell(row=i, column=2).value
        for i in range(3, formula_sheet.max_row)
        if formula_sheet.cell(row=i, column=1).value
    }
    # print(formula_dict)
    functions = {
        'ADD': lambda x, y: x + y,
        'MIN': lambda x, y: min(x, y),
        'MUL': lambda x, y: x * y
    }

    results = {'year': val_dict['year']}
    results_1 = {'year': val_dict_1['year']}
    for var in formula_dict.keys():
        formula = formula_dict.get(var).split('(')
        operands = formula[-1].strip(' ').strip(')').split(',')
        func = formula[0]
        oper_0 = operands[0]
        oper_1 = operands[-1] if len(operands[-1]) == 1 else operands[-1][-1]
        # print(oper_0)
        # print(oper_1)
        res = functions.get(func)(val_dict.get(oper_0), val_dict.get(oper_1))
        val_dict[var] = res
        results[var] = res
        res_1 = functions.get(func)(val_dict_1.get(oper_0), val_dict_1.get(oper_1))
        val_dict_1[var] = res_1
        results_1[var] = res_1
    # print(results)
    # print(results_1)
    append_excel(
        file_name='fintastic/exercise2.xlsx',
        results=results,
        results_1=results_1
    )


def test_xls():
    excel_file = openpyxl.load_workbook(r'fintastic/exercise2.xlsx')

    # del excel_file['Results']
    # excel_file.remove_sheet(res)
    # excel_file.remove_sheet('Results1')
    print(excel_file.sheetnames)

    # res_sheet = excel_file['Results']
    # res_dict = {
    #     res_sheet.cell(row=i, column=1).value:
    #         res_sheet.cell(row=i, column=2).value
    #     for i in range(1, res_sheet.max_row)
    #     if res_sheet.cell(row=i, column=1).value
    # }
    # print(res_dict)

    # excel_file.save(r'fintastic/exercise1.xlsx')


if __name__ == '__main__':
    proc_excel()
    # append_excel(file_name='fintastic/exercise1.xlsx', results={'int_1': 20, 'int_2': 15, 'final': 225})
    # test_xls()