import openpyxl


def append_excel(file_name, results):
    excel_file = openpyxl.load_workbook(fr'{file_name}')
    excel_file.create_sheet(title='Results', index=3)

    new_sheet = excel_file['Results']
    new_sheet.cell(row=1, column=1).value = 'Table 1'
    new_sheet.cell(row=2, column=1).value = 'Name'
    new_sheet.cell(row=2, column=2).value = 'Result'

    for row in results.items():
        new_sheet.append(row)

    excel_file.save(filename=file_name)


def proc_excel():
    excel_file = openpyxl.load_workbook(r'fintastic/exercise1.xlsx')
    values_sheet = excel_file['Values']
    formula_sheet = excel_file['Formula']
    val_dict = {
        values_sheet.cell(row=i, column=1).value:
        values_sheet.cell(row=i, column=2).value
        for i in range(3, values_sheet.max_row)
        if values_sheet.cell(row=i, column=1).value
    }
    formula_dict = {
        formula_sheet.cell(row=i, column=1).value:
        formula_sheet.cell(row=i, column=2).value
        for i in range(3, formula_sheet.max_row)
        if formula_sheet.cell(row=i, column=1).value
    }
    functions = {
        'ADD': lambda x, y: x + y,
        'MIN': lambda x, y: min(x, y),
        'MUL': lambda x, y: x * y
    }
    results = {}
    for var in formula_dict.keys():
        formula = formula_dict.get(var).split('(')
        operands = formula[-1].strip(' ').strip(')').split(',')
        func = formula[0]
        oper_0 = operands[0]
        oper_1 = operands[-1] if len(operands[-1]) == 1 else operands[-1][-1]
        res = functions.get(func)(val_dict.get(oper_0), val_dict.get(oper_1))
        val_dict[var] = res
        results[var] = res
    append_excel(
        file_name='fintastic/exercise1.xlsx',
        results=results
    )


def test_xls():
    excel_file = openpyxl.load_workbook(r'fintastic/exercise1.xlsx')

    # del excel_file['Results']
    # excel_file.remove_sheet(res)
    # excel_file.remove_sheet('Results1')
    print(excel_file.sheetnames)

    # res_sheet = excel_file['Results']
    # res_dict = {
    #     res_sheet.cell(row=i, column=1).value:
    #         res_sheet.cell(row=i, column=2).value
    #     for i in range(1, res_sheet.max_row)
    #     if res_sheet.cell(row=i, column=1).value
    # }
    # print(res_dict)

    # excel_file.save(r'fintastic/exercise1.xlsx')


if __name__ == '__main__':
    # proc_excel()
    # append_excel(file_name='fintastic/exercise1.xlsx', results={'int_1': 20, 'int_2': 15, 'final': 225})
    test_xls()